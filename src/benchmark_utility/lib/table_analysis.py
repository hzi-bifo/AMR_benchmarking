#!/usr/bin/python
import sys,os
# sys.path.append('../')
sys.path.insert(0, os.getcwd())
from src.amr_utility import name_utility,file_utility
import json
import pandas as pd
import numpy as np
import copy
from openpyxl import load_workbook
from src.benchmark_utility.lib.CombineResults import combine_data
import itertools
from scipy.stats import ttest_rel,ttest_ind

'''
This script organizes the performance for Supplementary materials, and further analysis on the results.
'''


def extract_info(level,s,fscore, f_all,output_path,step,tool_list,foldset,com_tool_list):
    main_meta,_=name_utility.GETname_main_meta(level)
    data = pd.read_csv(main_meta, index_col=0, dtype={'genome_id': object}, sep="\t")
    data = data[data['number'] != 0]  # drop the species with 0 in column 'number'.
    if f_all == False:#should not use it.
        data = data.loc[s, :]
    species_list=['Escherichia coli','Staphylococcus aureus','Salmonella enterica','Klebsiella pneumoniae','Pseudomonas aeruginosa',
                  'Acinetobacter baumannii','Streptococcus pneumoniae','Mycobacterium tuberculosis', 'Campylobacter jejuni',
                  'Enterococcus faecium','Neisseria gonorrhoeae']
    data=data.loc[species_list,:]
    df_species = data.index.tolist()
    antibiotics= data['modelling antibiotics'].tolist()

    # ------------------------------------------
    # Step 1.1 figuring out ML generally performance. and compare with ResFinder
    # ------------------------------------------
    # ------------------------------------------
    # # Step 1.2 figuring out the number of combinations that each tool performs better than ML baseline
    # # ------------------------------------------
    if step=='1':
        if com_tool_list==['Point-/ResFinder']:
            path_table_results2=output_path+ 'Results/other_figures_tables/ML_Com_resfinder_'+fscore+'.xlsx'
        elif com_tool_list==['ML Baseline (Majority)']:
            path_table_results2=output_path+ 'Results/other_figures_tables/ML_Com_MLbaseline_'+fscore+'.xlsx'
        else:
            print('Please add a new name manually at ./src/benchmark_utility/lib/table_analysis.py \
            if a new software except resfinder and ML baseline is chosed for a comparison with the rest.')
            exit(1)


        df1 = pd.DataFrame(index=species_list)
        file_utility.make_dir(os.path.dirname(path_table_results2))
        df1.to_excel(path_table_results2, sheet_name='introduction')
        score_list=['f1_macro', 'f1_positive', 'f1_negative', 'accuracy','clinical_f1_negative','clinical_precision_neg', 'clinical_recall_neg']
        for com_tool in com_tool_list:
            #each time count the cases the com_tool outperforms others.
            for eachfold in foldset:
                df_final=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software']+score_list)
                df_com=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software']+score_list)
                print('Compare with:',com_tool,eachfold,'-----')
                for species, antibiotics_selected in zip(df_species, antibiotics):
                    species_sub=[species]
                    tool_list_rest=copy.deepcopy(tool_list)
                    df_macro=combine_data(species_sub,level,'f1_macro',tool_list_rest,[eachfold],output_path)
                    df_acu=combine_data(species_sub,level,'accuracy',tool_list_rest,[eachfold],output_path)
                    df_neg=combine_data(species_sub,level,'f1_negative',tool_list_rest,[eachfold],output_path)
                    df_pos=combine_data(species_sub,level,'f1_positive',tool_list_rest,[eachfold],output_path)
                    df_cl_f1=combine_data(species_sub,level,'clinical_f1_negative',tool_list_rest,[eachfold],output_path)
                    df_cl_pre=combine_data(species_sub,level,'clinical_precision_neg',tool_list_rest,[eachfold],output_path)
                    df_cl_rec=combine_data(species_sub,level,'clinical_recall_neg',tool_list_rest,[eachfold],output_path)

                    df_macro['f1_negative']=df_neg['f1_negative']
                    df_macro['f1_positive']=df_pos['f1_positive']
                    df_macro['accuracy']=df_acu['accuracy']
                    df_macro['clinical_f1_negative']=df_cl_f1['clinical_f1_negative'].round(2)
                    df_macro['clinical_precision_neg']=df_cl_pre['clinical_precision_neg'].round(2)
                    df_macro['clinical_recall_neg']=df_cl_rec['clinical_recall_neg'].round(2)


                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])
                    df_macro = df_macro[['species', 'antibiotics', 'folds', 'software']+score_list]
                    df_final= pd.concat([df_final,df_macro])

                    # -----compare tool, based on fscore scores.
                    df_com_sub=combine_data(species_sub,level,fscore,[com_tool],[eachfold],output_path)
                    df_com= pd.concat([df_com,df_com_sub])


                df_com[fscore] = df_com[fscore] .astype(str)
                #### df_com['compare_'+fscore]=df_com[fscore].apply(lambda x:x.split('±')[0] if (len(x.split('±'))==2 ) else x)
                df_com['compare_'+fscore+'_mean']=df_com[fscore].apply(lambda x:x.split('±')[0])
                df_com['compare_'+fscore+'_std']=df_com[fscore].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else np.nan)
                df_com=df_com[['species', 'antibiotics','compare_'+fscore+'_mean','compare_'+fscore+'_std']]

                df_final[fscore] = df_final[fscore] .astype(str)

                df_final[fscore+'_mean']=df_final[fscore].apply(lambda x:x.split('±')[0])
                df_final[fscore+'_std']=df_final[fscore].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else np.nan)
                # -----------------------------------
                # ---make a comparison

                df_final=pd.merge(df_final, df_com, how="left", on=['species', 'antibiotics'])

                # ----------------------------------------


                df_final[fscore+'_mean'] = df_final[fscore+'_mean'] .astype(float)
                df_final[fscore+'_std'] = df_final[fscore+'_std'] .astype(float)
                df_final['compare_'+fscore+'_mean'] = df_final['compare_'+fscore+'_mean'] .astype(float).round(2)
                df_final['compare_'+fscore+'_std'] = df_final['compare_'+fscore+'_std'] .astype(float).round(2)

                wb = load_workbook(path_table_results2)
                ew = pd.ExcelWriter(path_table_results2)
                ew.book = wb
                df_final.to_excel(ew,sheet_name = (eachfold.split(' ')[0][0]+eachfold.split(' ')[1][0]+'_Comp_'+str(com_tool.translate(str.maketrans({'/': '', ' ': '_'})))))
                ew.save()
                #### Paired T test
                df_test=df_final[[fscore+'_mean', 'compare_'+fscore+'_mean']]
                df_test=df_test.fillna(0)
                print(df_test)
                mean1 = df_test[fscore+'_mean']
                mean2 = df_test['compare_'+fscore+'_mean']
                _,pvalue=ttest_rel(mean1, mean2,alternative='less') #
                print(pvalue)
                _,pvalue=ttest_rel(mean1, mean2,alternative='greater') #random
                print(pvalue)



    # ------------------------------------------
    # Step 2 figuring out which method performs best. And counting.
    # ------------------------------------------
    if step=='2':
        if tool_list==['Point-/ResFinder', 'Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']:

            if fscore=='f1_macro':
                path_table_results3_1=output_path+ 'Results/supplement_figures_tables/S6-1_software_winner_'+fscore+'.xlsx'
                path_table_results3_2=output_path+ 'Results/final_figures_tables/F3_results_heatmap_'+fscore+'.xlsx'
                path_table_results3_3=output_path+ 'Results/other_figures_tables/software_winner_'+fscore
            else: #clinical-oriented
                path_table_results3_1=output_path+ 'Results/other_figures_tables/software_winner_'+fscore+'.xlsx'
                path_table_results3_2=output_path+ 'Results/other_figures_tables/results_heatmap_'+fscore+'.xlsx'
        elif tool_list==['Point-/ResFinder', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover','Single-species-antibiotic Aytan-Aktug',
                   'Single-species multi-antibiotics Aytan-Aktug','Discrete databases multi-species model',
                'Concatenated databases mixed multi-species model', 'Concatenated databases leave-one-out multi-species model']:
            if fscore=='f1_macro':
                path_table_results3_1=output_path+ 'Results/supplement_figures_tables/S6-3_software_winner_multiModel_'+fscore+'.xlsx'
                path_table_results3_2=output_path+ 'Results/final_figures_tables/results_heatmap_multiModel_'+fscore+'.xlsx'
            else:#clinical-oriented
                path_table_results3_1=output_path+ 'Results/other_figures_tables/software_winner_multiModel_'+fscore+'.xlsx'
                path_table_results3_2=output_path+ 'Results/other_figures_tables/results_heatmap_multiModel_'+fscore+'.xlsx'
        else:
            print('Please add a new name manually at ./src/benchmark_utility/lib/table_analysis.py \
            if new software combinations are used for deciding winner or generate heatmap format excel.')
            exit(1)

        df1 = pd.DataFrame(index=species_list)
        file_utility.make_dir(os.path.dirname(path_table_results3_1))
        df1.to_excel(path_table_results3_1, sheet_name='introduction')

        #### each time count the cases the com_tool outperforms others.
        for eachfold in foldset:

            i=0
            for each_tool in tool_list:
                print(each_tool,eachfold,'-----')
                df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
                for species, antibiotics_selected in zip(df_species, antibiotics):

                    species_sub=[species]
                    df_macro=combine_data(species_sub,level,fscore,[each_tool],[eachfold],output_path)
                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])

                    df_macro[fscore] = df_macro[fscore].astype(str)
                    df_macro[each_tool]=df_macro[fscore].apply(lambda x:x.split('±')[0])
                    df_macro[each_tool+'_std']=df_macro[fscore].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else 10)
                    df_macro[each_tool] = df_macro[each_tool] .astype(float)
                    df_macro[each_tool+'_std'] = df_macro[each_tool+'_std'] .astype(float)

                    df_macro=df_macro[['species', 'antibiotics',each_tool,each_tool+'_std']]
                    df_final= pd.concat([df_final,df_macro])

                if i==0:
                    df_compare=df_final
                else:
                    df_compare=pd.merge(df_compare, df_final, how="left", on=['species', 'antibiotics'])
                i+=1


            df_compare=df_compare.round(2)
            df_std=df_compare[[x+'_std' for x in tool_list]]

            df_compare['max_'+fscore]=df_compare[tool_list].max(axis=1)
            a = df_compare[tool_list]
            df = a.eq(a.max(axis=1), axis=0)

            #considering of std
            for index, row in df.iterrows():

                winner=[]
                winner_std=[]
                for columnIndex, value in row.items():
                    # print(columnIndex,value, end="\t")
                    if value==True:
                        winner.append(columnIndex)
                        winner_std.append(df_std.loc[index,columnIndex+'_std'])
                if len(winner)>1: #more than two winner, check std

                    min_std = min(winner_std)
                    winner_index=[i for i, x in enumerate(winner_std) if x == min_std]
                    winner_filter=np.array(winner)[winner_index]
                    filter=list(set(winner) - set(winner_filter))
                    for each in filter:
                        row[each]=False
            df_compare=df_compare.replace({10: np.nan})
            df_compare['winner'] = df.mul(df.columns.to_series()).apply(','.join, axis=1).str.strip(',')
            df_compare=df_compare[['species', 'antibiotics']+tool_list+['max_'+fscore]+[x+'_std' for x in tool_list]+['winner' ]]
            df_compare.to_csv(path_table_results3_3+'_'+eachfold+'.csv', sep="\t") #for annotating heatmap

            wb = load_workbook(path_table_results3_1)
            ew = pd.ExcelWriter(path_table_results3_1)
            ew.book = wb
            df_compare.to_excel(ew,sheet_name = (eachfold))
            ew.save()

            #counting for each software the time it is the best, tied best had the coresponding portion of a 1.
            print('counting the perdentage of being best;')
            for each_tool in tool_list:
                print(each_tool,'-------------------------')
                count=0
                for each in df_compare['winner'].tolist():
                    if each_tool==each:
                        count+=1
                    elif each_tool in each:
                        # print(each)
                        winner_list=each.split(',')
                        winner_list = [x for x in winner_list if x != '']
                        # print(winner_list)
                        count+= (1/len(winner_list))
                print(count,count/len(df_compare['winner'].tolist()))
        #--------------------------------
        #mean +- std verson


        df1 = pd.DataFrame(index=species_list)
        file_utility.make_dir(os.path.dirname(path_table_results3_2))
        df1.to_excel(path_table_results3_2, sheet_name='introduction')
        # foldset=['random folds','phylo-tree-based folds','KMA-based folds']
        # tool_list=['Point-/ResFinder', 'Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']
            #each time count the cases the com_tool outperforms others.
        for eachfold in foldset:

            i=0
            for each_tool in tool_list:

                df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
                for species, antibiotics_selected in zip(df_species, antibiotics):
                    species_sub=[species]
                    df_macro=combine_data(species_sub,level,fscore,[each_tool],[eachfold],output_path)
                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])
                    df_macro[each_tool]=df_macro[fscore]
                    df_macro=df_macro[['species', 'antibiotics',each_tool]]
                    df_final= pd.concat([df_final,df_macro])

                if i==0:
                    df_compare=df_final
                else:
                    df_compare=pd.merge(df_compare, df_final, how="left", on=['species', 'antibiotics'])
                i+=1

            df_compare=df_compare[['species', 'antibiotics']+tool_list]


            with open('./data/AntiAcronym_dict.json') as f:
                map_acr = json.load(f)
            df_compare['antibiotics']=df_compare['antibiotics'].apply(lambda x: map_acr[x] )

            wb = load_workbook(path_table_results3_2)
            ew = pd.ExcelWriter(path_table_results3_2)
            ew.book = wb
            df_compare.to_excel(ew,sheet_name = (eachfold))
            ew.save()





    if step=='3':
        # paired T-test
        print('Now paired T-test')
        output=pd.DataFrame(columns=['folds','software','F1 diff', "p value"])
        Presults={}
        i_noDiff=0
        i_out=0
        for eachfold in foldset:
            print(eachfold)
            i=0
            for each_tool in tool_list:

                df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
                for species in  species_list:

                    species_sub=[species]
                    df_score=combine_data(species_sub,level,fscore,[each_tool],[eachfold],output_path)
                    df_score=df_score.reset_index()
                    df_score=df_score.drop(columns=['index'])

                    df_score[fscore] = df_score[fscore].astype(str)
                    df_score[each_tool]=df_score[fscore].apply(lambda x:x.split('±')[0]) #get mean if there is mean
                    df_score[each_tool] = df_score[each_tool] .astype(float)
                    df_score=df_score[['species', 'antibiotics',each_tool]]
                    df_final= pd.concat([df_final,df_score])

                if i==0:
                    df_compare=df_final
                else:
                    df_compare=pd.merge(df_compare, df_final, how="left", on=['species', 'antibiotics'])
                i+=1


            df_mean=df_compare[tool_list]

            # df_mean = df_mean.dropna()
            ### Feb 2023: maybe change the Nan to 0
            df_mean=df_mean.fillna(0)

            print('Paired T-test:')
            #T-test

            for each_com in list(itertools.combinations(tool_list, 2)):
                mean1 = df_mean[each_com[0]]
                mean2 = df_mean[each_com[1]]
                # result=ttest_rel(mean1, mean2) # two-tailed
                # pvalue = result[1]
                ###Feb 17th 2023, change to one-tailed test.
                # print(np.mean(mean1),np.mean(mean2))
                # Orders are based on Table 1 in the article.
                if eachfold=='Random folds':
                    order=[ 'Kover','PhenotypeSeeker','Point-/ResFinder', 'Seq2Geno2Pheno', 'Aytan-Aktug']
                elif eachfold=='Phylogeny-aware folds':
                    order=[ 'Point-/ResFinder','Kover','PhenotypeSeeker', 'Seq2Geno2Pheno', 'Aytan-Aktug']
                else:
                    order=[ 'Point-/ResFinder', 'Kover','PhenotypeSeeker','Seq2Geno2Pheno', 'Aytan-Aktug']

                # if np.mean(mean1) < np.mean(mean2):
                if order.index(each_com[0]) > order.index(each_com[1]):
                    _,pvalue=ttest_rel(mean1, mean2,alternative='less')
                else:
                    _,pvalue=ttest_rel(mean1, mean2,alternative='greater')
                print(each_com,pvalue)
                Presults[str( [each_com]+[eachfold])]=pvalue
                i_out+=1

                output.loc[i_out] = [eachfold, each_com,np.mean(mean1) - np.mean(mean2), pvalue]

                if pvalue>=0.05:
                    i_noDiff+=1
        print('No. can not reject null hypothesis:',i_noDiff)
        output.to_csv(output_path+ 'Results/supplement_figures_tables/S6-2_software_Pvalue_'+fscore+'_dropNan_1s_order.csv', sep="\t")
        # with open(output_path+ 'Results/supplement_figures_tables/S6-2_software_Pvalue_'+fscore+'.json', 'w') as f:
        #     json.dump(Presults, f)





    # if step=='4':
    #     # paired T-test, in the pool there are results from 4 ML methods and
    #     # 4 set of results of ResFinder (each set : the Resfinder results for the corresponding folds)
    #     print('Now paired T-test')
    #     pd.set_option('display.max_rows', 500)
    #     # pd.set_option('display.max_columns', 8)
    #     # # pd.set_option('display.width', 1000)
    #
    #     for eachfold in foldset:
    #         print(eachfold)
    #         i=0
    #         for each_tool in tool_list:
    #
    #             df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
    #             for species in  species_list:
    #
    #                 species_sub=[species]
    #                 df_score=combine_data(species_sub,level,fscore,[each_tool],[eachfold],output_path)
    #                 df_score=df_score.reset_index()
    #                 df_score=df_score.drop(columns=['index'])
    #
    #                 df_score[fscore] = df_score[fscore].astype(str)
    #                 df_score[each_tool]=df_score[fscore].apply(lambda x:x.split('±')[0]) #get mean if there is mean
    #                 df_score[each_tool] = df_score[each_tool] .astype(float)
    #                 df_score=df_score[['species', 'antibiotics',each_tool]]
    #                 df_final= pd.concat([df_final,df_score])
    #
    #             if i==0:
    #                 df_compare=df_final
    #             else:
    #                 df_compare=pd.merge(df_compare, df_final, how="left", on=['species', 'antibiotics'])
    #             i+=1
    #
    #
    #         df_mean=df_compare[['species', 'antibiotics']+tool_list]
    #         df_mean=df_mean.fillna(0)
    #         print(df_mean)
    #
    #
    #         table_eachfolds=pd.DataFrame(columns=['Point-/ResFinder','ML'])
    #         for each_com in ['Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']:
    #             table_eachfolds=table_eachfolds.rename({'ML':each_com}, axis=1)
    #             table_eachfolds=pd.concat([table_eachfolds, df_mean['Point-/ResFinder',each_com]], ignore_index=True)
    #
    #
    #         table_eachfolds.to_csv(output_path+ 'Results/other_figures_tables/S6-2_software_Pvalue_'+fscore+'.csv', sep="\t")

