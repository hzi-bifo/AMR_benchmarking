import amr_utility.name_utility
import amr_utility.load_data
import amr_utility.file_utility
import pickle
import pandas as pd
import numpy as np
import statistics
import copy
from openpyxl import load_workbook
from scipy.stats import f_oneway

"""This script organizes the performance 4 scores for Supplementary materials."""
'''
further anaysis on the results 
'''
def combine_data_res(species_list,fscore):
     # This function makes a matrix of all tools' results.
    df_plot = pd.DataFrame(columns=[fscore, 'species', 'software','antibiotics','folds'])
    columns_name=df_plot.columns.tolist()
    df_plot_sub = pd.DataFrame(columns=columns_name)
    # print(tool_list)
    # folds=['random folds','phylo-tree-based folds','KMA-based folds']

    for species in species_list:#'Point-/ResFinder':
            antibiotics, ID, Y = amr_utility.load_data.extract_info(species, False, 'loose')
            for anti in antibiotics:
                results_file='./benchmarking2_kma/resfinder/Results/summary/loose/'+str(species.replace(" ", "_"))+'.csv'
                results=pd.read_csv(results_file, header=0, index_col=0,sep="\t")
                score=results.loc[anti,fscore]
                df_plot_sub.loc['s'] = [score,species,'Point-/ResFinder',anti,'-']
                df_plot = df_plot.append(df_plot_sub, sort=False)

    return df_plot


def combine_data(species_list,fscore,tool_list,folds):

    # This function makes a matrix of all tools' results.
    df_plot = pd.DataFrame(columns=[fscore, 'species', 'software','antibiotics','folds'])
    columns_name=df_plot.columns.tolist()
    df_plot_sub = pd.DataFrame(columns=columns_name)
    # print(tool_list)
    # folds=['random folds','phylo-tree-based folds','KMA-based folds']
    # if 'Point-/ResFinder' in tool_list:
    #     for species in species_list:#'Point-/ResFinder':
    #         antibiotics, ID, Y = amr_utility.load_data.extract_info(species, False, 'loose')
    #         for anti in antibiotics:
    #             results_file='./benchmarking2_kma/resfinder/Results/summary/loose/'+str(species.replace(" ", "_"))+'.csv'
    #             results=pd.read_csv(results_file, header=0, index_col=0,sep="\t")
    #             score=results.loc[anti,fscore]
    #             df_plot_sub.loc['s'] = [score,species,'Point-/ResFinder',anti,'-']
    #             df_plot = df_plot.append(df_plot_sub, sort=False)

    # -----------------------------------
    for fold in folds:
        if fold=='phylo-tree-based folds':
            f_phylotree=True
            f_kma=False
        elif fold=='KMA-based folds':
            f_phylotree=False
            f_kma=True
        else:
            f_phylotree=False
            f_kma=False
        for species in species_list:
            if species !='Mycobacterium tuberculosis' or f_phylotree==False:
                antibiotics, ID, Y = amr_utility.load_data.extract_info(species, False, 'loose')
                for anti in antibiotics:

                    for tool in tool_list:
                        if tool=='Point-/ResFinder':
                            results_file,_ = amr_utility.name_utility.GETsave_name_final(fscore,species, f_kma, f_phylotree, 'resfinder')
                            if f_kma:
                                results_file='./resfinder_folds/'+results_file
                            elif f_phylotree:
                                results_file='./resfinder_folds/'+results_file
                            else:
                                results_file='./resfinder_folds/'+results_file

                            results=pd.read_csv(results_file + '.txt', header=0, index_col=0,sep="\t")

                            score=results.loc[anti,fscore]

                        if tool=='Aytan-Aktug':
                            results_file = amr_utility.name_utility.GETname_multi_bench_save_name_final(fscore,species, None,
                                                                                                                         'loose',
                                                                                                                         0.0,
                                                                                                                         0,
                                                                                                                         True,
                                                                                                                         False,
                                                                                                                         'f1_macro')
                            results_file='./benchmarking2_kma/'+results_file
                            if f_phylotree :
                                results_file=results_file+'_score_final_Tree.txt'
                                results=pd.read_csv(results_file, header=0, index_col=0,sep="\t")
                                score=results.loc[anti,fscore]
                            elif f_kma:
                                results_file=results_file+'_score_final.txt'
                                fscore_="weighted-"+fscore
                                results=pd.read_csv(results_file, header=0, index_col=0,sep="\t")
                                score=results.loc[anti,fscore_]
                            else:
                                results_file=results_file+ '_score_final_Random.txt'
                                results=pd.read_csv(results_file, header=0, index_col=0,sep="\t")
                                score=results.loc[anti,fscore]

                        # if tool=='KmerC':
                        #     if species !='Mycobacterium tuberculosis' :#no MT information.
                        #         _, results_file = amr_utility.name_utility.GETsave_name_final(fscore,species, f_kma, f_phylotree, '')
                        #         results_file='./patric_2022/'+results_file
                        #         results=pd.read_csv(results_file + '_SummeryBenchmarking.txt', header=0, index_col=0,sep="\t")
                        #         score=results.loc[anti,fscore]
                        #     else:
                        #         score=np.nan
                        if tool=='Seq2Geno2Pheno':
                            if species !='Mycobacterium tuberculosis':#no MT information.
                                _, results_file = amr_utility.name_utility.GETsave_name_final(fscore,species, f_kma, f_phylotree, '')
                                results_file='./seq2geno/'+results_file
                                results=pd.read_csv(results_file + '_SummeryBenchmarking.txt', header=0, index_col=0,sep="\t")
                                score=results.loc[anti,fscore]
                            else:

                                score=np.nan
                        if tool=='PhenotypeSeeker':
                            # if species !='Mycobacterium tuberculosis':
                            _, results_file = amr_utility.name_utility.GETsave_name_final(fscore,species, f_kma, f_phylotree, '')
                            if f_kma:
                                results_file='./PhenotypeSeeker_Nov08/'+results_file
                            elif f_phylotree:
                                results_file='./PhenotypeSeeker_tree/'+results_file
                            else:
                                results_file='./PhenotypeSeeker_random/'+results_file

                            results=pd.read_csv(results_file + '_SummeryBenchmarking.txt', header=0, index_col=0,sep="\t")
                            score=results.loc[anti,fscore]
                            # else:
                            #     score=np.nan
                        if tool=='Kover':

                            _, results_file = amr_utility.name_utility.GETsave_name_final(fscore,species, f_kma, f_phylotree, '')
                            if f_kma:
                                results_file='./kover/'+results_file
                            elif f_phylotree:
                                results_file='./kover_tree/'+results_file
                            else:
                                results_file='./kover_random/'+results_file
                            results=pd.read_csv(results_file + '_SummeryBenchmarking.txt', header=0, index_col=0,sep="\t")
                            score=results.loc[anti,fscore]

                        if tool=='Baseline (Majority)':

                            results_file,_ = amr_utility.name_utility.GETsave_name_final(fscore,species, f_kma, f_phylotree, 'majority')
                            if f_kma:
                                results_file='./majority/'+results_file
                            elif f_phylotree:
                                results_file='./majority/'+results_file
                            else:
                                results_file='./majority/'+results_file

                            results=pd.read_csv(results_file + '.txt', header=0, index_col=0,sep="\t")

                            score=results.loc[anti,fscore]
                        #[fscore, 'species', 'software','anti','folds']
                        df_plot_sub.loc['s'] = [score,species,tool,anti,fold]
                        df_plot = df_plot.append(df_plot_sub, sort=False)
    return df_plot

def extract_info(level,s, fscore, f_all,step):
    data = pd.read_csv('metadata/' + str(level) + '_Species_antibiotic_FineQuality.csv', index_col=0,
                       dtype={'genome_id': object}, sep="\t")
    data = data[data['number'] != 0]  # drop the species with 0 in column 'number'.
    if f_all == False:#should not use it.
        data = data.loc[s, :]
    species_list=['Escherichia coli','Staphylococcus aureus','Salmonella enterica','Klebsiella pneumoniae','Pseudomonas aeruginosa','Acinetobacter baumannii','Streptococcus pneumoniae','Mycobacterium tuberculosis', 'Campylobacter jejuni','Enterococcus faecium','Neisseria gonorrhoeae']
    data=data.loc[species_list,:]
    df_species = data.index.tolist()
    antibiotics= data['modelling antibiotics'].tolist()


    # initialize
    # f = open('log/results/cv_results.xlsx', 'w+')#The file is created if it does not exist, otherwise it is truncated.

    # folds=['phylo-tree-based folds','random folds']

    # ------------------------------------------
    # Step 1 figuring out ML generally performance. and compare with ResFinder
    # ------------------------------------------
    if step=='1':
        df1 = pd.DataFrame(index=species_list)
        df1.to_excel("log/results/cv_results_abstract1.xlsx", sheet_name='introduction')
        foldset=['random folds','phylo-tree-based folds','KMA-based folds']
        tool_list=[ 'Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']
        for eachfold in foldset:
            df_final=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
            df_resfinder=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
            print(df_species,'-----')
            for species, antibiotics_selected in zip(df_species, antibiotics):
                # summary= pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
                print(species)

                # _, ID, Y = amr_utility.load_data.extract_info(species, False, level)
                species_sub=[species]
                df_macro=combine_data(species_sub,'f1_macro',tool_list,[eachfold])
                df_acu=combine_data(species_sub,'accuracy',tool_list,[eachfold])
                df_neg=combine_data(species_sub,'f1_negative',tool_list,[eachfold])
                df_pos=combine_data(species_sub,'f1_positive',tool_list,[eachfold])
                # df_macro.insert(loc=0, column='species', value=[species] * (df_macro.shape[0]))
                df_macro['f1_negative']=df_neg['f1_negative']
                df_macro['f1_positive']=df_pos['f1_positive']
                df_macro['accuracy']=df_acu['accuracy']

                df_macro=df_macro.reset_index()
                df_macro=df_macro.drop(columns=['index'])
                df_macro = df_macro[['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy']]
                df_final= pd.concat([df_final,df_macro])
                # print(df_final)
            #
                df_resfinder_sub=combine_data(species_sub,'f1_macro',['Point-/ResFinder'],[eachfold])
                df_resfinder= pd.concat([df_resfinder,df_resfinder_sub])
            # #
            df_resfinder= df_resfinder.rename(columns={"f1_macro":"resfinder_f1_macro"})
            df_resfinder=df_resfinder[['species', 'antibiotics','folds',"resfinder_f1_macro"]]
            df_resfinder['resfinder_f1_macro'] = df_resfinder['resfinder_f1_macro'] .astype(str)
            df_resfinder['resfinder_f1_macro_mean']=df_resfinder['resfinder_f1_macro'].apply(lambda x:x.split('±')[0] if (len(x.split('±'))==2 ) else 0) #df_final['f1_macro']
            df_resfinder['resfinder_f1_macro_std']=df_resfinder['resfinder_f1_macro'].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else 10)


            # #macro f1 score > 0.9
            df_final['f1_macro'] = df_final['f1_macro'] .astype(str)
            # df_final.astype({'f1_macro': 'str'}).dtypes
            df_final['ML_f1macro_mean']=df_final['f1_macro'].apply(lambda x:x.split('±')[0] if (len(x.split('±'))==2 ) else 0) #df_final['f1_macro']
            df_final['ML_f1macro_std']=df_final['f1_macro'].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else 10)
            # -----------------------------------
            # ---make a comparison with Resfinder

            # print(df_resfinder)
            df_final=pd.merge(df_final, df_resfinder, how="left", on=['species', 'antibiotics','folds'])

            # ----------------------------------------


            df_final['ML_f1macro_mean'] = df_final['ML_f1macro_mean'] .astype(float)
            df_final['ML_f1macro_std'] = df_final['ML_f1macro_std'] .astype(float)
            df_final['resfinder_f1_macro_mean'] = df_final['resfinder_f1_macro_mean'] .astype(float)
            df_final['resfinder_f1_macro_std'] = df_final['resfinder_f1_macro_std'] .astype(float)

            # df_final.astype({'f1macro_mean': 'float','f1macro_std': 'float'}).dtypes
            print(df_final)
            wb = load_workbook('log/results/cv_results_abstract1.xlsx')
            ew = pd.ExcelWriter('log/results/cv_results_abstract1.xlsx')
            ew.book = wb
            df_final.to_excel(ew,sheet_name = eachfold)
            ew.save()

    # ------------------------------------------
    # # Step 2 figuring out the number of combinations that each tool performs best(or tied first ). for ML baseline
    # # ------------------------------------------
    if step=='2':
        df1 = pd.DataFrame(index=species_list)
        df1.to_excel("log/results/cv_results_abstract2.xlsx", sheet_name='introduction')
        foldset=['random folds','phylo-tree-based folds','KMA-based folds']
        tool_list=['Point-/ResFinder','Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']
        for com_tool in ['Baseline (Majority)']:
            #each time count the cases the com_tool outperforms others.
            for eachfold in foldset:
                df_final=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
                df_com=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
                print(com_tool,eachfold,'-----')
                for species, antibiotics_selected in zip(df_species, antibiotics):
                    # summary= pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
                    print(species)

                    # _, ID, Y = amr_utility.load_data.extract_info(species, False, level)
                    species_sub=[species]

                    tool_list_rest=copy.deepcopy(tool_list)
                    # tool_list_rest.remove(com_tool)
                    df_macro=combine_data(species_sub,'f1_macro',tool_list_rest,[eachfold])
                    df_acu=combine_data(species_sub,'accuracy',tool_list_rest,[eachfold])
                    df_neg=combine_data(species_sub,'f1_negative',tool_list_rest,[eachfold])
                    df_pos=combine_data(species_sub,'f1_positive',tool_list_rest,[eachfold])
                    # df_macro.insert(loc=0, column='species', value=[species] * (df_macro.shape[0]))
                    df_macro['f1_negative']=df_neg['f1_negative']
                    df_macro['f1_positive']=df_pos['f1_positive']
                    df_macro['accuracy']=df_acu['accuracy']

                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])
                    df_macro = df_macro[['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy']]
                    df_final= pd.concat([df_final,df_macro])
                    # print(df_final)


                    # -----compare tool, now only compare f1_macro scores.
                    df_com_sub=combine_data(species_sub,'f1_macro',[com_tool],[eachfold])
                    df_com= pd.concat([df_com,df_com_sub])
                #
                # df_com= df_com.rename(columns={"f1_macro":"compare_f1_macro"})
                df_com['f1_macro'] = df_com['f1_macro'] .astype(str)
                df_com['compare_f1_macro']=df_com['f1_macro'].apply(lambda x:x.split('±')[0] if (len(x.split('±'))==2 ) else 0)

                df_com=df_com[['species', 'antibiotics',"compare_f1_macro"]]
                # #macro f1 score > 0.9
                df_final['f1_macro'] = df_final['f1_macro'] .astype(str)
                # df_final.astype({'f1_macro': 'str'}).dtypes
                df_final['f1macro_mean']=df_final['f1_macro'].apply(lambda x:x.split('±')[0] if (len(x.split('±'))==2 ) else 0) #df_final['f1_macro']
                df_final['f1macro_std']=df_final['f1_macro'].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else 10)
                # -----------------------------------
                # ---make a comparison with Resfinder

                # print(df_resfinder)
                df_final=pd.merge(df_final, df_com, how="left", on=['species', 'antibiotics'])

                # ----------------------------------------


                df_final['f1macro_mean'] = df_final['f1macro_mean'] .astype(float)
                df_final['f1macro_std'] = df_final['f1macro_std'] .astype(float)
                # df_final.astype({'f1macro_mean': 'float','f1macro_std': 'float'}).dtypes
                print(df_final)

                wb = load_workbook('log/results/cv_results_abstract2.xlsx')
                ew = pd.ExcelWriter('log/results/cv_results_abstract2.xlsx')
                ew.book = wb
                # if com_tool=='Point-/ResFinder':
                #     df_final.to_excel(ew,sheet_name = (eachfold+'_resfinder'))
                # else:
                df_final.to_excel(ew,sheet_name = (eachfold+'_'+com_tool))
                ew.save()



    # ------------------------------------------
    # Step 3 figuring out which ML performs best.
    # ------------------------------------------
    if step=='3':
        df1 = pd.DataFrame(index=species_list)
        df1.to_excel("log/results/cv_results_abstract3.xlsx", sheet_name='introduction')
        foldset=['random folds','phylo-tree-based folds','KMA-based folds']
        tool_list=['Point-/ResFinder', 'Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']
            #each time count the cases the com_tool outperforms others.
        for eachfold in foldset:
            # # df_resfinder=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
            # df_resfinder=combine_data(species_sub,'f1_macro',[each_tool],[eachfold])
            # df_resfinder=df_macro.reset_index()
            # df_macro=df_macro.drop(columns=['index'])

            i=0
            for each_tool in tool_list:
                print(each_tool,eachfold,'-----')
                df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
                for species, antibiotics_selected in zip(df_species, antibiotics):
                    # summary= pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
                    print(species)

                    # _, ID, Y = amr_utility.load_data.extract_info(species, False, level)
                    species_sub=[species]
                    df_macro=combine_data(species_sub,'f1_macro',[each_tool],[eachfold])
                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])
                    # df_macro = df_macro[['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy']]
                    # df_macro=df_macro[['species', 'antibiotics','f1_macro']]
                    df_macro['f1_macro'] = df_macro['f1_macro'].astype(str)
                    df_macro[each_tool]=df_macro['f1_macro'].apply(lambda x:x.split('±')[0] if (len(x.split('±'))==2 ) else 0) #df_final['f1_macro']
                    df_macro[each_tool+'_std']=df_macro['f1_macro'].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else 10)
                    df_macro[each_tool] = df_macro[each_tool] .astype(float)
                    df_macro[each_tool+'_std'] = df_macro[each_tool+'_std'] .astype(float)
                    df_macro=df_macro[['species', 'antibiotics',each_tool,each_tool+'_std']]
                    # df_macro= df_macro.rename(columns={"f1_macro":each_tool+"_f1_macro"})
                    df_final= pd.concat([df_final,df_macro])
                    # print(df_final)
                if i==0:
                    df_compare=df_final
                else:
                    df_compare=pd.merge(df_compare, df_final, how="left", on=['species', 'antibiotics'])
                i+=1

            print(df_compare)
            df_std=df_compare[['Point-/ResFinder_std','Aytan-Aktug_std','Seq2Geno2Pheno_std',\
                                                       'PhenotypeSeeker_std','Kover_std']]
            df_compare['max_f1_macro']=df_compare[['Point-/ResFinder','Aytan-Aktug','Seq2Geno2Pheno',\
                                                       'PhenotypeSeeker','Kover']].max(axis=1)
            a = df_compare[['Point-/ResFinder','Aytan-Aktug','Seq2Geno2Pheno',\
                                                       'PhenotypeSeeker','Kover']]
            df = a.eq(a.max(axis=1), axis=0)
            print(df)
            #todo add, considering of std
            for index, row in df.iterrows():

                winner=[]
                winner_std=[]
                for columnIndex, value in row.items():
                    # print(columnIndex,value, end="\t")
                    if value==True:
                        winner.append(columnIndex)
                        winner_std.append(df_std.loc[index,columnIndex+'_std'])
                if len(winner)>1: #more than two winner, check std
                    print(winner_std)
                    min_std = min(winner_std)
                    winner_index=[i for i, x in enumerate(winner_std) if x == min_std]
                    winner_filter=np.array(winner)[winner_index]
                    filter=list(set(winner) - set(winner_filter))
                    print(winner)
                    print(winner_filter)
                    print(filter)
                    for each in filter:
                        row[each]=False

            print(df)
            df_compare['winner'] = df.mul(df.columns.to_series()).apply(','.join, axis=1).str.strip(',')
            df_compare=df_compare[['species', 'antibiotics','Point-/ResFinder','Aytan-Aktug','Seq2Geno2Pheno',\
                                                       'PhenotypeSeeker','Kover', 'max_f1_macro','Point-/ResFinder_std','Aytan-Aktug_std','Seq2Geno2Pheno_std',\
                                                       'PhenotypeSeeker_std','Kover_std','winner' ]]
            # df_resfinder['max_tool']=df_resfinder[['resfinder_f1_macro','Aytan-Aktug_f1_macro','Seq2Geno2Pheno_f1_macro',\
            #                                            'PhenotypeSeeker_f1_macro','Kover_f1_macro',]].idxmax(axis=1)
            # df_resfinder['winner'] = df_resfinder['winner'].astype(str)
            # df_resfinder['max_tool']=df_resfinder['winner'].apply(lambda x:x.split('_')[0] if len(x.split(','))==0 else x)
            wb = load_workbook('log/results/cv_results_abstract3.xlsx')
            ew = pd.ExcelWriter('log/results/cv_results_abstract3.xlsx')
            ew.book = wb
            df_compare.to_excel(ew,sheet_name = (eachfold))
            ew.save()
        #--------------------------------
        #mean +- std verson


        df1 = pd.DataFrame(index=species_list)
        df1.to_excel("log/results/cv_results_abstract3_2.xlsx", sheet_name='introduction')
        foldset=['random folds','phylo-tree-based folds','KMA-based folds']
        tool_list=['Point-/ResFinder', 'Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']
            #each time count the cases the com_tool outperforms others.
        for eachfold in foldset:
            # # df_resfinder=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
            # df_resfinder=combine_data(species_sub,'f1_macro',[each_tool],[eachfold])
            # df_resfinder=df_macro.reset_index()
            # df_macro=df_macro.drop(columns=['index'])

            i=0
            for each_tool in tool_list:
                print(each_tool,eachfold,'-----')
                df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
                for species, antibiotics_selected in zip(df_species, antibiotics):
                    # summary= pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])
                    print(species)

                    # _, ID, Y = amr_utility.load_data.extract_info(species, False, level)
                    species_sub=[species]
                    df_macro=combine_data(species_sub,'f1_macro',[each_tool],[eachfold])
                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])
                    # df_macro = df_macro[['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy']]
                    # df_macro=df_macro[['species', 'antibiotics','f1_macro']]
                    # df_macro['f1_macro'] = df_macro['f1_macro'].astype(str)
                    df_macro[each_tool]=df_macro['f1_macro']
                    # df_macro[each_tool]=df_macro['f1_macro'].apply(lambda x:x.split('±')[0] if (len(x.split('±'))==2 ) else 0) #df_final['f1_macro']
                    # df_macro[each_tool+'_std']=df_macro['f1_macro'].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else 10)
                    # df_macro[each_tool] = df_macro[each_tool] .astype(float)
                    # df_macro[each_tool+'_std'] = df_macro[each_tool+'_std'] .astype(float)
                    df_macro=df_macro[['species', 'antibiotics',each_tool]]
                    # df_macro= df_macro.rename(columns={"f1_macro":each_tool+"_f1_macro"})
                    df_final= pd.concat([df_final,df_macro])
                    # print(df_final)
                if i==0:
                    df_compare=df_final
                else:
                    df_compare=pd.merge(df_compare, df_final, how="left", on=['species', 'antibiotics'])
                i+=1

            print(df_compare)
            df_compare=df_compare[['species', 'antibiotics','Point-/ResFinder','Aytan-Aktug','Seq2Geno2Pheno',\
                                                       'PhenotypeSeeker','Kover']]

            with open('./src/AntiAcronym_dict.pkl', 'rb') as f:
                map_acr = pickle.load(f)
            df_compare['antibiotics']=df_compare['antibiotics'].apply(lambda x: map_acr[x] )
            # df_resfinder['max_tool']=df_resfinder[['resfinder_f1_macro','Aytan-Aktug_f1_macro','Seq2Geno2Pheno_f1_macro',\
            #                                            'PhenotypeSeeker_f1_macro','Kover_f1_macro',]].idxmax(axis=1)
            # df_resfinder['winner'] = df_resfinder['winner'].astype(str)
            # df_resfinder['max_tool']=df_resfinder['winner'].apply(lambda x:x.split('_')[0] if len(x.split(','))==0 else x)
            wb = load_workbook('log/results/cv_results_abstract3_2.xlsx')
            ew = pd.ExcelWriter('log/results/cv_results_abstract3_2.xlsx')
            ew.book = wb
            df_compare.to_excel(ew,sheet_name = (eachfold))
            ew.save()



    if step=='4':
    #todo, check for each species, the best software.
        df1 = pd.DataFrame(index=species_list)
        df1.to_excel("log/results/cv_results_abstract4.xlsx", sheet_name='introduction')
        foldset=['random folds','phylo-tree-based folds','KMA-based folds']
        tool_list=['Point-/ResFinder','Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']
            #each time count the cases the com_tool outperforms others.


        for species, antibiotics_selected in zip(df_species, antibiotics):
            print(species)
            for eachfold in foldset:
                for each_tool in tool_list: #todo
                    print(each_tool,eachfold,'-----')
                    df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
                    # for species, antibiotics_selected in zip(df_species, antibiotics):
                        # summary= pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'f1_macro', 'f1_positive', 'f1_negative', 'accuracy'])


                    # _, ID, Y = amr_utility.load_data.extract_info(species, False, level)
                    species_sub=[species]
                    df_macro=combine_data(species_sub,'f1_macro',[each_tool],[eachfold])
                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])
                    # df_macro = df_macro[['species', 'antibiotics', 'folds', 'software','f1_macro', 'f1_positive', 'f1_negative', 'accuracy']]
                    # df_macro=df_macro[['species', 'antibiotics','f1_macro']]
                    df_macro['f1_macro'] = df_macro['f1_macro'].astype(str)
                    df_macro[each_tool]=df_macro['f1_macro'].apply(lambda x:x.split('±')[0] if (len(x.split('±'))==2 ) else 0) #df_final['f1_macro']
                    df_macro[each_tool] = df_macro[each_tool] .astype(float)
                    df_macro=df_macro[['species', 'antibiotics',each_tool]]
                    # df_macro= df_macro.rename(columns={"f1_macro":each_tool+"_f1_macro"})
                    df_final= pd.concat([df_final,df_macro])
                    # print(df_final)


            df_final['max_f1_macro']=df_final[['Point-/ResFinder','Aytan-Aktug','Seq2Geno2Pheno',\
                                                       'PhenotypeSeeker','Kover',]].max(axis=1)
            a = df_final[['Point-/ResFinder','Aytan-Aktug','Seq2Geno2Pheno',\
                                                       'PhenotypeSeeker','Kover',]]
            df = a.eq(a.max(axis=1), axis=0)

            df_final['winner'] = df.mul(df.columns.to_series())\
                       .apply(','.join, axis=1)\
                       .str.strip(',')

            # df_resfinder['max_tool']=df_resfinder[['resfinder_f1_macro','Aytan-Aktug_f1_macro','Seq2Geno2Pheno_f1_macro',\
            #                                            'PhenotypeSeeker_f1_macro','Kover_f1_macro',]].idxmax(axis=1)
            # df_resfinder['winner'] = df_resfinder['winner'].astype(str)
            # df_resfinder['max_tool']=df_resfinder['winner'].apply(lambda x:x.split('_')[0] if len(x.split(','))==0 else x)
            wb = load_workbook('log/results/cv_results_abstract4.xlsx')
            ew = pd.ExcelWriter('log/results/cv_results_abstract4.xlsx')
            ew.book = wb
            df_final.to_excel(ew,sheet_name = (species))
            ew.save()
